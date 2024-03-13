import os
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.conf import settings

def fit_column_width(ws):
    for col in ws.iter_cols():
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass  # In case of NoneType or other data types that don't support len()
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width
    return ws


def hide_columns(ws):
    cols_to_be_hidden = [1, 2, 5, 6, 8, 10, 12] + np.arange(14,18).tolist() + np.arange(20,41).tolist() + np.arange(42,51).tolist() + np.arange(52,60).tolist() + np.arange(62,80).tolist()

    for col_index in cols_to_be_hidden:
        col_letter = get_column_letter(col_index)
        ws.column_dimensions[col_letter].hidden = True
    return ws

def apply_border(ws):
    # Define the border style

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
                cell.border = thin_border
    return ws

def color_courses(ws):
                
    # Define colors (add more colors as needed)
    colors = [
        'FFF2CC',  # Light Yellow
        'E2EFDA',  # Light Green
        'DEEBF7',  # Light Blue
        'FCE4D6',  # Light Orange
        'D9EAD3',  # Very Light Green
        'D6E3F8',  # Very Light Blue
        'FCF2F2',  # Very Light Red
        'EAD1DC',  # Light Purple
        'F4CCCC',  # Soft Red
        'D9D2E9',  # Soft Purple
    ]
    color_map = {}
    
    # Starting from row 2 to skip header, apply color based on unique 'Course Name'
    for row in range(2, ws.max_row + 1):
        course_name = ws[f'AO{row}'].value # Assuming 'Course Name' is in column AO
        if course_name not in color_map:
            # Assign a new color from the list, and rotate the list
            color_map[course_name], colors = colors[0], colors[1:] + colors[:1]
        fill = PatternFill(start_color = color_map[course_name], end_color=color_map[course_name], fill_type="solid")
        for cell in ws[row]:
            cell.fill = fill
    return ws

def style_num_students(ws):
    
    def format_row(row):
        for cell in row:
            cell.fill = PatternFill(start_color="004B96", end_color="004B96", fill_type="solid")
            cell.font = Font(color="FFFFFF",bold=True)
            cell.alignment = Alignment(horizontal="center")
    
    format_row(ws[1])
    format_row(ws[ws.max_row])
    
    ws.column_dimensions["B"].width = 18
    
    ws.column_dimensions["A"].width = 60
    
    return ws
    

def apply_styling(sheet):
    if sheet.title == "num_studying":
        ws = style_num_students(sheet)
        ws = apply_border(ws)
    elif sheet.title == "all":
        ws = fit_column_width(sheet)
        ws = apply_border(ws)
        ws = hide_columns(ws)
    else:
        ws = fit_column_width(sheet)
        ws = apply_border(ws)
        ws = hide_columns(ws)
        ws = color_courses(ws)
        
    return ws


def generate_studying_status(excel_file):
    
    # Define columns
    date_columns = ["Date Of Birth", "English Test Date", "Proposed Start Date", "Proposed End Date",
                    "Actual Start Date", "Actual End Date", "Visa Effective Date", "Visa End Date",
                    "Visa Non Grant Action Date", "COE Created Date", "COE Last Updated"]
    str_columns = ["Passport Number", "Mobile", "Phone", "Student Address Line 1", "Visa Grant Number"]
    str_col_dict = {key: "str" for key in str_columns}

    # Load and process the Excel file
    df = pd.read_excel(excel_file, header=3, parse_dates=date_columns, dtype=str_col_dict)

    # # Using `header=None` to not treat any row as headers
    # # Access the content of cell A3, which is at index position [2, 0] (3rd row, 1st column)
    # created = pd.read_excel(excel_file, header=None, nrows=3).iloc[2,0]

    
    studying = df[df["COE Status"] == "Studying"]
    studying = studying.sort_values(by=["Course Name", "Proposed Start Date"])

    course_counts = studying["Course Name"].value_counts().reset_index()
    course_counts.columns = [f"Studying students as of readable_date_time", "CoE"]
    totals = pd.DataFrame({f"Studying students as of readable_date_time": ['Total'], 'CoE': [course_counts['CoE'].sum()]})
    course_counts = pd.concat([course_counts, totals], ignore_index=True)

    approved = df[(df["COE Status"] == "Approved") |  (df["COE Status"] == "Visa Granted")]
    approved = approved.sort_values(by=["Proposed Start Date", "Course Name"])

    finished = df[df["COE Status"] == "Finished"]
    finished = finished.sort_values(by="Proposed End Date")

    current_month = datetime.now().month
    next_month = (datetime.now() + timedelta(days=30)).month
    current_year = datetime.now().year
    finishing = df[(df["Proposed End Date"].dt.year == current_year) & ((df["Proposed End Date"].dt.month == current_month) | (df["Proposed End Date"].dt.month == next_month)) ]
    finishing = finishing.sort_values(by=["Proposed End Date"])

    dfs = {"num_studying": course_counts, "all": df, "studying": studying, "approved": approved, "finishing": finishing, "finished": finished}
   
    # Assuming final_file_path is determined after processing
    # readable_date_time = datetime.now().strftime('%d.%m.%Y') # Placeholder for actual logic
    final_file_path = os.path.join(settings.MEDIA_ROOT, "Studying Status Report.xlsx")
    
    # Ensure the MEDIA_ROOT directory exists
    if not os.path.exists(settings.MEDIA_ROOT):
        os.makedirs(settings.MEDIA_ROOT)

    # Create a Pandas Excel writer using ExcelWriter
    with pd.ExcelWriter(final_file_path, engine='openpyxl') as writer:
        # DataFrame processing and saving logic here
        # Loop through the dictionary and write each DataFrame to a different sheet
        for sheet_name, df in dfs.items():
            if sheet_name != "num_studying":
                for date_col in date_columns:
                    df[date_col] = df[date_col].dt.date
            df.to_excel(writer, sheet_name=sheet_name, index = False)

    # Load the workbook and apply styling
    wb = load_workbook(final_file_path)
    # Assume apply_styling function is defined elsewhere or included here
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        apply_styling(sheet)  # Your styling function applied here

    # Save the styled workbook
    wb.save(final_file_path)

    return final_file_path

